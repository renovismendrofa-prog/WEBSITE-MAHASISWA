from dotenv import load_dotenv
load_dotenv()
from flask import Flask, render_template, request, redirect, url_for, session, flash, Response
import sqlite3, csv, json, io, os, time, re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

app = Flask(__name__)
app.secret_key = "ganti_dengan_secret_key_random"

DB = "database.db"

# ── Admin credentials ───────────────────────────────────────────────────
# Password admin disimpan di sini, bisa diganti lewat fitur lupa password
ADMIN_USERS = {"admin": os.environ.get("ADMIN_PASSWORD", "admin123")}

# Email admin (untuk terima OTP dan laporan)
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "")

# ── OTP store (in-memory, hilang kalau server restart) ──────────────────
import random, hashlib
OTP_STORE = {}   # { "admin": {"otp": "123456", "expires": timestamp} }

def generate_otp():
    return str(random.randint(100000, 999999))

def store_otp(username, otp):
    OTP_STORE[username] = {"otp": otp, "expires": time.time() + 300}  # 5 menit

def verify_otp(username, otp_input):
    data = OTP_STORE.get(username)
    if not data:
        return False, "OTP tidak ditemukan. Minta OTP baru."
    if time.time() > data["expires"]:
        OTP_STORE.pop(username, None)
        return False, "OTP sudah kadaluarsa (5 menit). Minta OTP baru."
    if data["otp"] != otp_input.strip():
        return False, "OTP salah."
    OTP_STORE.pop(username, None)
    return True, None

# ── Email (Gmail SMTP) config ────────────────────────────────────────────
# Isi dengan email Gmail kamu dan App Password (BUKAN password Gmail biasa).
# Cara buat App Password: myaccount.google.com/apppasswords (wajib 2FA aktif dulu)
GMAIL_ADDRESS      = os.environ.get("GMAIL_ADDRESS", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")

def send_email(to_address, subject, body, html_body=None):
    """Kirim email lewat Gmail SMTP. Support plain text dan HTML."""
    if not to_address:
        return False, "Penerima belum punya alamat email."
    if not GMAIL_ADDRESS or not GMAIL_APP_PASSWORD:
        return False, "Email pengirim belum dikonfigurasi. Set environment variable GMAIL_ADDRESS dan GMAIL_APP_PASSWORD."

    msg = MIMEMultipart("alternative")
    msg["From"]    = GMAIL_ADDRESS
    msg["To"]      = to_address
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))
    if html_body:
        msg.attach(MIMEText(html_body, "html"))

    try:
        with smtplib.SMTP("smtp.gmail.com", 587, timeout=15) as server:
            server.starttls()
            server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_ADDRESS, to_address, msg.as_string())
        return True, None
    except smtplib.SMTPAuthenticationError:
        return False, "Login Gmail gagal. Periksa GMAIL_ADDRESS dan App Password."
    except Exception as e:
        return False, f"Gagal mengirim email: {e}"

def send_laporan_email(rows):
    """Kirim laporan semua data mahasiswa ke email admin."""
    if not ADMIN_EMAIL:
        return False, "Email admin belum dikonfigurasi. Set environment variable ADMIN_EMAIL."

    # Build tabel HTML
    rows_html = ""
    for i, r in enumerate(rows, 1):
        ipk_color = "#3B6D11" if r["ipk"] >= 3.0 else "#854F0B" if r["ipk"] >= 2.5 else "#A32D2D"
        rows_html += f"""
        <tr style="background:{'#f9f9f9' if i%2==0 else '#ffffff'}">
          <td style="padding:8px 12px;border-bottom:1px solid #e5e7eb">{i}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e5e7eb;font-family:monospace;font-size:12px">{r['nim']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e5e7eb;font-weight:500">{r['nama']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e5e7eb">{r.get('email') or '—'}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e5e7eb">{r['jurusan']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e5e7eb;color:{ipk_color};font-weight:600">{r['ipk']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e5e7eb">{r['status']}</td>
        </tr>"""

    avg_ipk = round(sum(r["ipk"] for r in rows) / len(rows), 2) if rows else 0
    aktif   = sum(1 for r in rows if r["status"] == "Aktif")

    html_body = f"""
    <div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;max-width:900px;margin:0 auto">
      <div style="background:#185FA5;padding:20px 28px;border-radius:10px 10px 0 0">
        <h1 style="color:#fff;margin:0;font-size:20px">📊 Laporan Data Mahasiswa</h1>
        <p style="color:#bfdbfe;margin:6px 0 0;font-size:13px">Dikirim otomatis dari Dashboard Academic Terpadu</p>
      </div>
      <div style="background:#f8f9fa;padding:16px 28px;border:1px solid #e5e7eb;border-top:none">
        <div style="display:flex;gap:24px">
          <div><span style="font-size:11px;color:#6b7280;text-transform:uppercase;letter-spacing:.05em">Total</span><br><strong style="font-size:22px">{len(rows)}</strong></div>
          <div><span style="font-size:11px;color:#6b7280;text-transform:uppercase;letter-spacing:.05em">Rata-rata IPK</span><br><strong style="font-size:22px">{avg_ipk}</strong></div>
          <div><span style="font-size:11px;color:#6b7280;text-transform:uppercase;letter-spacing:.05em">Aktif</span><br><strong style="font-size:22px">{aktif}</strong></div>
        </div>
      </div>
      <table style="width:100%;border-collapse:collapse;border:1px solid #e5e7eb;border-top:none">
        <thead>
          <tr style="background:#f1f5f9">
            <th style="padding:10px 12px;text-align:left;font-size:11px;color:#6b7280;text-transform:uppercase;border-bottom:2px solid #e5e7eb">#</th>
            <th style="padding:10px 12px;text-align:left;font-size:11px;color:#6b7280;text-transform:uppercase;border-bottom:2px solid #e5e7eb">NIM</th>
            <th style="padding:10px 12px;text-align:left;font-size:11px;color:#6b7280;text-transform:uppercase;border-bottom:2px solid #e5e7eb">Nama</th>
            <th style="padding:10px 12px;text-align:left;font-size:11px;color:#6b7280;text-transform:uppercase;border-bottom:2px solid #e5e7eb">Email</th>
            <th style="padding:10px 12px;text-align:left;font-size:11px;color:#6b7280;text-transform:uppercase;border-bottom:2px solid #e5e7eb">Jurusan</th>
            <th style="padding:10px 12px;text-align:left;font-size:11px;color:#6b7280;text-transform:uppercase;border-bottom:2px solid #e5e7eb">IPK</th>
            <th style="padding:10px 12px;text-align:left;font-size:11px;color:#6b7280;text-transform:uppercase;border-bottom:2px solid #e5e7eb">Status</th>
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>
      <p style="font-size:11px;color:#9ca3af;padding:12px 0">Dikirim pada {time.strftime('%d %B %Y, %H:%M')} WIB</p>
    </div>"""

    plain = f"Laporan Data Mahasiswa\nTotal: {len(rows)} | Rata-rata IPK: {avg_ipk} | Aktif: {aktif}\n\n"
    for r in rows:
        plain += f"{r['nim']} | {r['nama']} | {r['jurusan']} | IPK {r['ipk']} | {r['status']}\n"

    return send_email(ADMIN_EMAIL, f"Laporan Data Mahasiswa — {len(rows)} mahasiswa", plain, html_body)

# ── DB helpers ─────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS mahasiswa (
                id      INTEGER PRIMARY KEY AUTOINCREMENT,
                nim     TEXT UNIQUE NOT NULL,
                nama    TEXT NOT NULL,
                email   TEXT DEFAULT '',
                ipk     REAL NOT NULL,
                jurusan TEXT NOT NULL,
                status  TEXT NOT NULL DEFAULT 'Aktif'
            )
        """)
        # Migrasi: tambah kolom email kalau DB lama belum punya
        cols = [r[1] for r in conn.execute("PRAGMA table_info(mahasiswa)").fetchall()]
        if "email" not in cols:
            conn.execute("ALTER TABLE mahasiswa ADD COLUMN email TEXT DEFAULT ''")

        if conn.execute("SELECT COUNT(*) FROM mahasiswa").fetchone()[0] == 0:
            seed = [
                ("101123456001","Ahmad Fauzi",    "ahmad.fauzi@student.ac.id",    3.75,"Informatika",      "Aktif"),
                ("101123456002","Budi Santoso",   "budi.santoso@student.ac.id",   3.20,"Sistem Informasi", "Aktif"),
                ("101123456003","Citra Dewi",     "citra.dewi@student.ac.id",     2.85,"Teknik Komputer",  "Cuti"),
                ("101123456004","Dian Pratama",   "dian.pratama@student.ac.id",   3.90,"Informatika",      "Aktif"),
                ("101123456005","Eka Rahmawati",  "eka.rahmawati@student.ac.id",  2.10,"Sistem Informasi", "Keluar"),
            ]
            conn.executemany(
                "INSERT INTO mahasiswa (nim,nama,email,ipk,jurusan,status) VALUES (?,?,?,?,?,?)", seed
            )

# ── Sorting algorithms ─────────────────────────────────────────────────
def bubble_sort(data, key, reverse=False):
    arr = list(data)
    n = len(arr)
    for i in range(n):
        for j in range(n - i - 1):
            a, b = arr[j][key], arr[j+1][key]
            if (a > b) if not reverse else (a < b):
                arr[j], arr[j+1] = arr[j+1], arr[j]
    return arr

def selection_sort(data, key, reverse=False):
    arr = list(data)
    n = len(arr)
    for i in range(n):
        idx = i
        for j in range(i+1, n):
            a, b = arr[j][key], arr[idx][key]
            if (a < b) if not reverse else (a > b):
                idx = j
        arr[i], arr[idx] = arr[idx], arr[i]
    return arr

def shell_sort(data, key, reverse=False):
    arr = list(data)
    n, gap = len(arr), len(arr) // 2
    while gap > 0:
        for i in range(gap, n):
            temp = arr[i]
            j = i
            while j >= gap and ((arr[j-gap][key] > temp[key]) if not reverse else (arr[j-gap][key] < temp[key])):
                arr[j] = arr[j-gap]
                j -= gap
            arr[j] = temp
        gap //= 2
    return arr

def merge_sort(data, key, reverse=False):
    arr = list(data)
    if len(arr) <= 1:
        return arr
    mid = len(arr) // 2
    left  = merge_sort(arr[:mid], key, reverse)
    right = merge_sort(arr[mid:], key, reverse)
    result, i, j = [], 0, 0
    while i < len(left) and j < len(right):
        a, b = left[i][key], right[j][key]
        if (a <= b) if not reverse else (a >= b):
            result.append(left[i]);  i += 1
        else:
            result.append(right[j]); j += 1
    result.extend(left[i:]); result.extend(right[j:])
    return result

SORT_FN = {
    "bubble_nim_asc":     lambda d: bubble_sort(d,    "nim", False),
    "bubble_nim_desc":    lambda d: bubble_sort(d,    "nim", True),
    "selection_nim_asc":  lambda d: selection_sort(d, "nim", False),
    "selection_nim_desc": lambda d: selection_sort(d, "nim", True),
    "shell_ipk_asc":      lambda d: shell_sort(d,     "ipk", False),
    "shell_ipk_desc":     lambda d: shell_sort(d,     "ipk", True),
    "merge_ipk_asc":      lambda d: merge_sort(d,     "ipk", False),
    "merge_ipk_desc":     lambda d: merge_sort(d,     "ipk", True),
}

# ── Search ─────────────────────────────────────────────────────────────
def linear_search(data, query):
    q = query.lower()
    return [r for r in data if
            q in r["nama"].lower() or
            q in r["nim"] or
            q in (r.get("email") or "").lower()]

def binary_search(data, nim):
    arr = sorted(data, key=lambda x: x["nim"])
    lo, hi = 0, len(arr) - 1
    while lo <= hi:
        mid = (lo + hi) // 2
        if arr[mid]["nim"] == nim:   return [arr[mid]]
        elif arr[mid]["nim"] < nim:  lo = mid + 1
        else:                        hi = mid - 1
    return []

# ── Auth decorators ──────────────────────────────────────────────────────
def login_required(f):
    """Boleh diakses oleh admin ATAU mahasiswa."""
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper

def admin_required(f):
    """Hanya boleh diakses oleh admin."""
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Akses ditolak. Halaman ini khusus admin.", "error")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return wrapper

# ── Validation helpers ─────────────────────────────────────────────────
EMAIL_RE = re.compile(r'^[a-zA-Z0-9_.+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')

def validate_email(email):
    if not email:
        return True   # email opsional
    return bool(EMAIL_RE.match(email))

# ── Routes: Auth ─────────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        u = request.form.get("username","").strip()
        p = request.form.get("password","").strip()

        # 1. Cek dulu apakah login sebagai admin
        if ADMIN_USERS.get(u) == p:
            session["user"] = u
            session["role"] = "admin"
            return redirect(url_for("index"))

        # 2. Cek login sebagai mahasiswa: username = NIM, password = 6 digit terakhir NIM
        if re.fullmatch(r"\d{12}", u) and len(p) == 6 and u.endswith(p):
            with get_db() as conn:
                row = conn.execute("SELECT * FROM mahasiswa WHERE nim=?", (u,)).fetchone()
            if row:
                session["user"]    = u
                session["role"]    = "mahasiswa"
                session["mhs_id"]  = row["id"]
                return redirect(url_for("profil"))

        flash("Username atau password salah.")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ── Lupa Password ──────────────────────────────────────────────────────────
@app.route("/lupa-password", methods=["GET","POST"])
def lupa_password():
    if request.method == "POST":
        username = request.form.get("username","").strip()
        if username not in ADMIN_USERS:
            flash("Username admin tidak ditemukan.", "error")
            return render_template("lupa_password.html", step="request")

        otp = generate_otp()
        store_otp(username, otp)

        # Cek apakah ADMIN_EMAIL sudah dikonfigurasi
        if "isi_email_admin" in ADMIN_EMAIL:
            flash("Email admin belum dikonfigurasi di server (ADMIN_EMAIL).", "error")
            return render_template("lupa_password.html", step="request")

        html_otp = f"""
        <div style="font-family:-apple-system,sans-serif;max-width:420px;margin:0 auto">
          <div style="background:#185FA5;padding:20px 28px;border-radius:10px 10px 0 0">
            <h2 style="color:#fff;margin:0;font-size:18px">🔐 Kode OTP Reset Password</h2>
          </div>
          <div style="background:#fff;padding:24px 28px;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 10px 10px">
            <p style="color:#374151;margin:0 0 16px">Kamu meminta reset password untuk akun <strong>{username}</strong>.</p>
            <div style="background:#f1f5f9;border-radius:8px;padding:20px;text-align:center;margin-bottom:16px">
              <p style="margin:0 0 6px;font-size:12px;color:#6b7280;text-transform:uppercase;letter-spacing:.05em">Kode OTP kamu</p>
              <p style="margin:0;font-size:36px;font-weight:700;letter-spacing:12px;color:#185FA5;font-family:monospace">{otp}</p>
            </div>
            <p style="color:#6b7280;font-size:13px;margin:0">Kode ini berlaku selama <strong>5 menit</strong>. Jangan bagikan ke siapapun.</p>
          </div>
        </div>"""

        ok, err = send_email(ADMIN_EMAIL, "Kode OTP Reset Password Dashboard Academic", f"Kode OTP kamu: {otp} (berlaku 5 menit)", html_otp)
        if not ok:
            flash(f"Gagal kirim OTP: {err}", "error")
            return render_template("lupa_password.html", step="request")

        session["otp_username"] = username
        flash(f"OTP telah dikirim ke email admin. Berlaku 5 menit.", "success")
        return redirect(url_for("verifikasi_otp"))

    return render_template("lupa_password.html", step="request")

@app.route("/verifikasi-otp", methods=["GET","POST"])
def verifikasi_otp():
    username = session.get("otp_username")
    if not username:
        return redirect(url_for("lupa_password"))

    if request.method == "POST":
        otp_input = request.form.get("otp","").strip()
        ok, err = verify_otp(username, otp_input)
        if not ok:
            flash(err, "error")
            return render_template("lupa_password.html", step="otp", username=username)
        session["otp_verified"] = True
        return redirect(url_for("reset_password"))

    return render_template("lupa_password.html", step="otp", username=username)

@app.route("/reset-password", methods=["GET","POST"])
def reset_password():
    username = session.get("otp_username")
    if not username or not session.get("otp_verified"):
        return redirect(url_for("lupa_password"))

    if request.method == "POST":
        pw1 = request.form.get("password","").strip()
        pw2 = request.form.get("confirm","").strip()

        if len(pw1) < 6:
            flash("Password minimal 6 karakter.", "error")
            return render_template("lupa_password.html", step="reset")
        if pw1 != pw2:
            flash("Password dan konfirmasi tidak cocok.", "error")
            return render_template("lupa_password.html", step="reset")

        ADMIN_USERS[username] = pw1
        session.pop("otp_username", None)
        session.pop("otp_verified", None)
        flash("Password berhasil direset! Silakan login.", "success")
        return redirect(url_for("login"))

    return render_template("lupa_password.html", step="reset")

# ── Kirim Laporan ──────────────────────────────────────────────────────────
@app.route("/kirim-laporan", methods=["POST"])
@admin_required
def kirim_laporan():
    with get_db() as conn:
        rows = [dict(r) for r in conn.execute("SELECT * FROM mahasiswa ORDER BY nim").fetchall()]

    if not rows:
        flash("Tidak ada data mahasiswa untuk dikirim.", "error")
        return redirect(url_for("index"))

    ok, err = send_laporan_email(rows)
    if ok:
        flash(f"Laporan {len(rows)} data mahasiswa berhasil dikirim ke {ADMIN_EMAIL}.", "success")
    else:
        flash(f"Gagal kirim laporan: {err}", "error")
    return redirect(url_for("index"))

# ── Routes: Admin dashboard ───────────────────────────────────────────────
@app.route("/", methods=["GET"])
@login_required
def index():
    # Mahasiswa diarahkan ke halaman profil, bukan dashboard admin
    if session.get("role") == "mahasiswa":
        return redirect(url_for("profil"))

    sort_key    = request.args.get("sort", "bubble_nim_asc")
    search_q    = request.args.get("q", "").strip()
    search_mode = request.args.get("mode", "linear")

    with get_db() as conn:
        rows = [dict(r) for r in conn.execute("SELECT * FROM mahasiswa").fetchall()]

    t0   = time.perf_counter()
    fn   = SORT_FN.get(sort_key, SORT_FN["bubble_nim_asc"])
    rows = fn(rows)
    elapsed = round((time.perf_counter() - t0) * 1000, 4)

    sort_labels = {
        "bubble_nim_asc":"Bubble Sort (NIM ↑)","bubble_nim_desc":"Bubble Sort (NIM ↓)",
        "selection_nim_asc":"Selection Sort (NIM ↑)","selection_nim_desc":"Selection Sort (NIM ↓)",
        "shell_ipk_asc":"Shell Sort (IPK ↑)","shell_ipk_desc":"Shell Sort (IPK ↓)",
        "merge_ipk_asc":"Merge Sort (IPK ↑)","merge_ipk_desc":"Merge Sort (IPK ↓)",
    }
    complexity = {"bubble":"O(n²)","selection":"O(n²)","shell":"O(n log² n)","merge":"O(n log n)"}
    algo    = sort_key.split("_")[0]
    log_msg = f"{sort_labels.get(sort_key, sort_key)} selesai dalam {elapsed} ms · {complexity.get(algo,'')}"

    if search_q:
        if search_mode == "binary" and re.fullmatch(r"\d{12}", search_q):
            rows = binary_search(rows, search_q)
        else:
            rows = linear_search(rows, search_q)

    all_rows = [dict(r) for r in get_db().execute("SELECT * FROM mahasiswa").fetchall()]
    total   = len(all_rows)
    avg_ipk = round(sum(r["ipk"] for r in all_rows) / total, 2) if total else 0
    aktif   = sum(1 for r in all_rows if r["status"] == "Aktif")
    cuti    = sum(1 for r in all_rows if r["status"] == "Cuti")
    keluar  = sum(1 for r in all_rows if r["status"] == "Keluar")

    return render_template("index.html",
        rows=rows, log_msg=log_msg,
        sort_key=sort_key, search_q=search_q, search_mode=search_mode,
        total=total, avg_ipk=avg_ipk, aktif=aktif, cuti=cuti, keluar=keluar
    )

@app.route("/tambah", methods=["GET","POST"])
@admin_required
def tambah():
    if request.method == "POST":
        nim     = request.form.get("nim","").strip()
        nama    = request.form.get("nama","").strip()
        email   = request.form.get("email","").strip().lower()
        ipk     = request.form.get("ipk","").strip()
        jurusan = request.form.get("jurusan","").strip()
        status  = request.form.get("status","Aktif")

        errors = []
        if not re.fullmatch(r"\d{12}", nim):
            errors.append("NIM harus tepat 12 digit angka.")
        if not nama:
            errors.append("Nama tidak boleh kosong.")
        if email and not validate_email(email):
            errors.append("Format email tidak valid.")
        try:
            ipk_val = float(ipk)
            if not (0 <= ipk_val <= 4): raise ValueError
        except ValueError:
            errors.append("IPK harus angka antara 0.00–4.00.")
        if not jurusan:
            errors.append("Jurusan tidak boleh kosong.")

        if errors:
            return render_template("tambah.html", errors=errors,
                nim=nim, nama=nama, email=email, ipk=ipk, jurusan=jurusan, status=status)
        try:
            with get_db() as conn:
                conn.execute(
                    "INSERT INTO mahasiswa (nim,nama,email,ipk,jurusan,status) VALUES (?,?,?,?,?,?)",
                    (nim, nama, email, float(ipk), jurusan, status)
                )
            flash("Data berhasil ditambahkan.", "success")
            return redirect(url_for("index"))
        except sqlite3.IntegrityError:
            return render_template("tambah.html",
                errors=["NIM sudah terdaftar di database."],
                nim=nim, nama=nama, email=email, ipk=ipk, jurusan=jurusan, status=status)

    return render_template("tambah.html", errors=[])

@app.route("/edit/<int:id>", methods=["GET","POST"])
@admin_required
def edit(id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM mahasiswa WHERE id=?", (id,)).fetchone()
    if not row:
        flash("Data tidak ditemukan.")
        return redirect(url_for("index"))

    if request.method == "POST":
        nama    = request.form.get("nama","").strip()
        email   = request.form.get("email","").strip().lower()
        ipk     = request.form.get("ipk","").strip()
        jurusan = request.form.get("jurusan","").strip()
        status  = request.form.get("status","Aktif")

        errors = []
        if not nama:   errors.append("Nama tidak boleh kosong.")
        if email and not validate_email(email):
            errors.append("Format email tidak valid.")
        try:
            ipk_val = float(ipk)
            if not (0 <= ipk_val <= 4): raise ValueError
        except ValueError:
            errors.append("IPK harus angka antara 0.00–4.00.")
        if not jurusan: errors.append("Jurusan tidak boleh kosong.")

        if errors:
            return render_template("tambah.html", errors=errors,
                nim=row["nim"], nama=nama, email=email, ipk=ipk,
                jurusan=jurusan, status=status, edit_id=id)

        with get_db() as conn:
            conn.execute(
                "UPDATE mahasiswa SET nama=?,email=?,ipk=?,jurusan=?,status=? WHERE id=?",
                (nama, email, float(ipk), jurusan, status, id)
            )
        flash("Data berhasil diperbarui.", "success")
        return redirect(url_for("index"))

    return render_template("tambah.html", errors=[],
        nim=row["nim"], nama=row["nama"], email=row["email"] or "",
        ipk=row["ipk"], jurusan=row["jurusan"], status=row["status"], edit_id=id)

@app.route("/hapus/<int:id>")
@admin_required
def hapus(id):
    with get_db() as conn:
        conn.execute("DELETE FROM mahasiswa WHERE id=?", (id,))
    flash("Data berhasil dihapus.", "success")
    return redirect(url_for("index"))

@app.route("/kirim-email/<int:id>", methods=["GET","POST"])
@admin_required
def kirim_email(id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM mahasiswa WHERE id=?", (id,)).fetchone()
    if not row:
        flash("Data mahasiswa tidak ditemukan.", "error")
        return redirect(url_for("index"))

    mhs = dict(row)

    if request.method == "POST":
        subject = request.form.get("subject","").strip()
        body    = request.form.get("body","").strip()

        if not mhs.get("email"):
            flash(f"{mhs['nama']} belum memiliki alamat email.", "error")
            return redirect(url_for("index"))
        if not subject or not body:
            return render_template("kirim_email.html", mhs=mhs,
                errors=["Subjek dan isi pesan tidak boleh kosong."],
                subject=subject, body=body)

        ok, err = send_email(mhs["email"], subject, body)
        if ok:
            flash(f"Email berhasil dikirim ke {mhs['nama']} ({mhs['email']}).", "success")
            return redirect(url_for("index"))
        else:
            return render_template("kirim_email.html", mhs=mhs,
                errors=[err], subject=subject, body=body)

    default_subject = "Informasi Akademik"
    default_body = (
        f"Halo {mhs['nama']},\n\n"
        f"Pesan ini dikirim melalui Dashboard Academic Terpadu.\n\n"
        f"Salam,\nAdmin Akademik"
    )
    return render_template("kirim_email.html", mhs=mhs, errors=[],
        subject=default_subject, body=default_body)

@app.route("/export")
@admin_required
def export():
    with get_db() as conn:
        rows = conn.execute("SELECT nim,nama,email,ipk,jurusan,status FROM mahasiswa").fetchall()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["nim","nama","email","ipk","jurusan","status"])
    for r in rows:
        w.writerow(list(r))
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition":"attachment;filename=data_mahasiswa.csv"})

@app.route("/import", methods=["POST"])
@admin_required
def bulk_import():
    f = request.files.get("file")
    if not f:
        flash("Pilih file terlebih dahulu.", "error")
        return redirect(url_for("index"))

    fname = f.filename.lower()
    inserted = skipped = 0

    try:
        if fname.endswith(".csv") or fname.endswith(".txt"):
            content = f.read().decode("utf-8")
            # deteksi delimiter otomatis (koma, titik koma, atau tab)
            sample = content[:1024]
            if sample.count(";") > sample.count(","):
                delim = ";"
            elif sample.count("\t") > sample.count(","):
                delim = "\t"
            else:
                delim = ","
            rows = list(csv.DictReader(io.StringIO(content), delimiter=delim))

        elif fname.endswith(".json"):
            rows = json.loads(f.read().decode("utf-8"))

        elif fname.endswith(".xlsx"):
            if not XLSX_AVAILABLE:
                flash("Server belum mendukung file Excel. Install openpyxl terlebih dahulu.", "error")
                return redirect(url_for("index"))
            wb = load_workbook(filename=io.BytesIO(f.read()), read_only=True, data_only=True)
            ws = wb.active
            data_iter = ws.iter_rows(values_only=True)
            headers = [str(h).strip().lower() if h else "" for h in next(data_iter)]
            rows = []
            for raw_row in data_iter:
                if all(c is None for c in raw_row):
                    continue
                rows.append({headers[i]: raw_row[i] for i in range(len(headers)) if i < len(raw_row)})

        else:
            flash("Format file harus .csv, .json, .xlsx, atau .txt", "error")
            return redirect(url_for("index"))

        with get_db() as conn:
            for r in rows:
                nim     = str(r.get("nim","") or "").strip()
                nama    = str(r.get("nama","") or "").strip()
                email   = str(r.get("email","") or "").strip().lower()
                jurusan = str(r.get("jurusan","") or "").strip()
                status  = str(r.get("status","Aktif") or "Aktif").strip()
                try:
                    ipk = float(r.get("ipk", 0) or 0)
                except (ValueError, TypeError):
                    skipped += 1; continue
                if not re.fullmatch(r"\d{12}", nim):
                    skipped += 1; continue
                if email and not validate_email(email):
                    email = ""
                try:
                    conn.execute(
                        "INSERT INTO mahasiswa (nim,nama,email,ipk,jurusan,status) VALUES (?,?,?,?,?,?)",
                        (nim, nama, email, ipk, jurusan, status)
                    )
                    inserted += 1
                except sqlite3.IntegrityError:
                    skipped += 1

        flash(f"Import selesai: {inserted} data ditambahkan, {skipped} dilewati.", "success")
    except Exception as e:
        flash(f"Gagal memproses file: {e}", "error")

    return redirect(url_for("index"))

# ── Routes: Mahasiswa portal ───────────────────────────────────────────────
@app.route("/profil", methods=["GET"])
@login_required
def profil():
    if session.get("role") != "mahasiswa":
        return redirect(url_for("index"))

    with get_db() as conn:
        me = conn.execute("SELECT * FROM mahasiswa WHERE id=?", (session["mhs_id"],)).fetchone()
        all_rows = [dict(r) for r in conn.execute("SELECT * FROM mahasiswa").fetchall()]

    if not me:
        session.clear()
        flash("Data mahasiswa tidak ditemukan. Silakan login kembali.")
        return redirect(url_for("login"))

    sort_key = request.args.get("sort", "bubble_nim_asc")
    fn = SORT_FN.get(sort_key, SORT_FN["bubble_nim_asc"])
    all_rows = fn(all_rows)

    total   = len(all_rows)
    avg_ipk = round(sum(r["ipk"] for r in all_rows) / total, 2) if total else 0

    return render_template("profil.html", me=dict(me), rows=all_rows,
        sort_key=sort_key, total=total, avg_ipk=avg_ipk)

@app.route("/profil/edit", methods=["GET","POST"])
@login_required
def profil_edit():
    if session.get("role") != "mahasiswa":
        return redirect(url_for("index"))

    with get_db() as conn:
        me = conn.execute("SELECT * FROM mahasiswa WHERE id=?", (session["mhs_id"],)).fetchone()
    if not me:
        session.clear()
        return redirect(url_for("login"))

    if request.method == "POST":
        email   = request.form.get("email","").strip().lower()
        jurusan = request.form.get("jurusan","").strip()

        errors = []
        if email and not validate_email(email):
            errors.append("Format email tidak valid.")
        if not jurusan:
            errors.append("Jurusan tidak boleh kosong.")

        if errors:
            return render_template("profil_edit.html", errors=errors,
                me=dict(me), email=email, jurusan=jurusan)

        with get_db() as conn:
            conn.execute(
                "UPDATE mahasiswa SET email=?, jurusan=? WHERE id=?",
                (email, jurusan, session["mhs_id"])
            )
        flash("Profil berhasil diperbarui.", "success")
        return redirect(url_for("profil"))

    return render_template("profil_edit.html", errors=[],
        me=dict(me), email=me["email"] or "", jurusan=me["jurusan"])

if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
